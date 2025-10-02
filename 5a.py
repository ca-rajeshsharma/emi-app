import streamlit as st
import openpyxl
from openpyxl.styles import Font
import math

def calculate_emi(principal, annual_rate, months):
    monthly_rate = annual_rate / (12 * 100)
    emi = principal * monthly_rate * math.pow(1 + monthly_rate, months) / (math.pow(1 + monthly_rate, months) - 1)
    return round(emi, 2)

def generate_detailed_emi_excel(party_name, principal, annual_rate, months):
    emi = calculate_emi(principal, annual_rate, months)
    monthly_rate = annual_rate / (12 * 100)
    total_payment = round(emi * months, 2)
    total_interest = round(total_payment - principal, 2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EMI Summary"

    summary_headers = ["Party Name", "Principal Amount", "Rate of Interest (%)", "Period (Months)", "Monthly EMI", "Total Payment", "Total Interest"]
    summary_values = [party_name, principal, annual_rate, months, emi, total_payment, total_interest]

    for col, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for col, value in enumerate(summary_values, start=1):
        ws.cell(row=2, column=col, value=value)

    ws_amort = wb.create_sheet(title="Amortization Schedule")
    amort_headers = ["Month", "EMI", "Principal Paid", "Interest Paid", "Remaining Balance"]
    for col, header in enumerate(amort_headers, start=1):
        cell = ws_amort.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    balance = principal
    for month in range(1, months + 1):
        interest = round(balance * monthly_rate, 2)
        principal_paid = round(emi - interest, 2)
        balance = round(balance - principal_paid, 2)
        ws_amort.append([month, emi, principal_paid, interest, balance])

    return wb

# 🎯 Streamlit UI
st.title("EMI Calculator with Excel Export")

party_name = st.text_input("Enter Party Name")
principal = st.number_input("Enter Principal Loan Amount", min_value=0.0)
annual_rate = st.number_input("Enter Annual Interest Rate (%)", min_value=0.0)
months = st.number_input("Enter Loan Period (Months)", min_value=1, step=1)

if st.button("Generate EMI Report"):
    if party_name and principal and annual_rate and months:
        wb = generate_detailed_emi_excel(party_name, principal, annual_rate, months)
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        st.download_button(
            label="Download EMI Excel Report",
            data=buffer.getvalue(),
            file_name="Detailed_EMI_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("Please fill in all fields.")
